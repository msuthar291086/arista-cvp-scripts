import json
from cvprac.cvp_client import CvpClient
import pprint
import ssl
import urllib3
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
import time
import datetime
import getpass

print "\nEnter CVP login details:-\n"

CVP_HOST = raw_input("CVP IP:")
CVP_USER = raw_input("Username:")
CVP_PW = getpass.getpass("Password: ")

ssl._create_default_https_context = ssl._create_unverified_context
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

client = CvpClient()
client.connect([CVP_HOST], CVP_USER, CVP_PW, protocol='https')

# Create a workbook and add a worksheet.
wb = Workbook()
sheet = wb.create_sheet(title = 'CVP_Inventory_Sheet', index = 0)
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 'Switch Hostname'
ws['B1'] = 'Switch MGMT IP'
ws['C1'] = 'EOS_Version'
ws['D1'] = 'Model'
ws['E1'] = 'Serial_Number'
ws['F1'] = 'MAC_Address'

sheet['A1'].font = Font(sz = 12, bold = True)
sheet['B1'].font = Font(sz = 12, bold = True)
sheet['C1'].font = Font(sz = 12, bold = True)
sheet['D1'].font = Font(sz = 12, bold = True)
sheet['E1'].font = Font(sz = 12, bold = True)
sheet['F1'].font = Font(sz = 12, bold = True)

sheet.column_dimensions['A'].width = 36
sheet.column_dimensions['B'].width = 15
sheet.column_dimensions['C'].width = 12
sheet.column_dimensions['D'].width = 17
sheet.column_dimensions['E'].width = 14
sheet.column_dimensions['F'].width = 16

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

ws.cell(row=1, column=1).border = thin_border
ws.cell(row=1, column=2).border = thin_border
ws.cell(row=1, column=3).border = thin_border
ws.cell(row=1, column=4).border = thin_border
ws.cell(row=1, column=5).border = thin_border
ws.cell(row=1, column=6).border = thin_border

cvp_info = client.get('/cvpInfo/getCvpInfo.do')

print "################################################################################"
print "CVP Version is %s" %(cvp_info['version'])

cvp_inventory = client.get('/inventory/devices')

no_of_devices = len(cvp_inventory)

x = 2
y = 2
z = 1
print "Total Number of Devices in CVP: "+str(no_of_devices)

for i in range(no_of_devices):
	sheet['A' + str(x)].value = cvp_inventory[i]["hostname"]
	sheet['B' + str(x)].value = cvp_inventory[i]["ipAddress"]
	sheet['C' + str(x)].value = cvp_inventory[i]["version"]
	sheet['D' + str(x)].value = cvp_inventory[i]["modelName"]
	sheet['E' + str(x)].value = cvp_inventory[i]["serialNumber"]
	sheet['F' + str(x)].value = cvp_inventory[i]["systemMacAddress"]
	ws.cell(row=y, column=z).border = thin_border
	ws.cell(row=y, column=z+1).border = thin_border
	ws.cell(row=y, column=z+2).border = thin_border
	ws.cell(row=y, column=z+3).border = thin_border
	ws.cell(row=y, column=z+4).border = thin_border
	ws.cell(row=y, column=z+5).border = thin_border
	x+=1
	y+=1


print "################################################################################"

localtime = datetime.datetime.now().strftime("%d-%m-%Y-%H-%M")

output_filename = "CVP_Device_Inventory_List-%s.xlsx" %(localtime)

ws.cell(row=3, column=2).border = thin_border

wb.save(output_filename)
